import openpyxl
import numpy as np
import pandas as pd
import streamlit as st

st.set_page_config(layout = "wide", initial_sidebar_state = 'expanded')

def ex2py(fecha):
    if pd.notnull(fecha):
        fecha = str(fecha)[:10]
    else:
        fecha = None
    return fecha

st.title('Informes Técnicos Favorables')

with st.sidebar: 
    excel  = st.file_uploader('Subir el reporte generado por el sistema:')
    if excel is not None:
        data     = pd.read_excel(excel, sheet_name = 'ITF-Concluidos')
        columnas = data.columns.tolist()
        cols_sel = []
        for i in range(len(columnas)):
            if st.checkbox(columnas[i], key = 'col_' + str(i)):
                cols_sel.append(columnas[i])

        
if excel is not None:
    if len(cols_sel) != 0: 
        st.dataframe(data[cols_sel], hide_index = True, use_container_width = True)
    if st.button('Generar excel arreglado', use_container_width = True):
        date1 = data['FCHA_EXPDNTE'].apply(ex2py)
        date2 = data['FECHA_EMISION_OFI'].apply(ex2py)
        date3 = []
        for i in range(date1.shape[0]):
            if date1.iat[i] != None and date2.iat[i] != None:
                date3.append(np.busday_count(date1.iat[i], date2.iat[i]))
            else:
                date3.append(None)
        date3 = pd.DataFrame(date3, columns = ['DIAS_HABILES'])
        date4 = pd.concat([data[cols_sel], date1, date2, date3], axis = 1)
        date4.to_excel('Actualizado.xlsx', index = False, engine = 'xlsxwriter')

        plantilla = openpyxl.load_workbook('Plantilla.xlsx')
        plan_hoja = plantilla['Sheet1']
        actualiza = openpyxl.load_workbook('Actualizado.xlsx')
        actu_hoja = actualiza['Sheet1']
        for i in range(plan_hoja.max_row - 1):
            for j in range(6):
                plan_hoja.cell(row = i + 2, column = j + 1).value = actu_hoja.cell(row = i + 2, column = j + 1).value
        plantilla.save('Arreglado.xlsx')

        st.download_button(
            label = "Descargar excel arreglado", data = open("Arreglado.xlsx", "rb").read(),
            file_name = "Arreglado.xlsx", use_container_width = True
        )        
